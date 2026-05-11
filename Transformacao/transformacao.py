import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import ast
import re
from pathlib import Path

path_atual = Path.cwd()
# Ler os arquivos
print("📂 Lendo arquivos...")
modelo = pd.read_excel(path_atual / 'modelo.xlsx', header=None)
modelo2 = pd.read_excel(path_atual / 'output' / 'resultado.xlsx')

print("\n✅ Arquivos carregados com sucesso!")

# Função para converter string de lista para lista real
def parse_list(s):
    """Converte string de lista para lista real"""
    if pd.isna(s):
        return []
    if isinstance(s, list):
        return s
    try:
        return ast.literal_eval(s)
    except:
        return []

# Função para limpar os dias (garantir formato padrão)
def normalizar_dia(dia):
    """Normaliza o nome do dia"""
    dias_map = {
        'seg': 'Segunda',
        'ter': 'Terça',
        'qua': 'Quarta',
        'qui': 'Quinta',
        'sex': 'Sexta',
        'sab': 'Sábado',
        'dom': 'Domingo',
        'segunda': 'Segunda',
        'terça': 'Terça',
        'quarta': 'Quarta',
        'quinta': 'Quinta',
        'sexta': 'Sexta',
        'sábado': 'Sábado',
        'domingo': 'Domingo',
    }
    return dias_map.get(dia.lower().strip(), dia)

# Extrair dados do modelo2
print("\n📊 Processando dados do modelo2...")
resultados = []

for idx, row in modelo2.iterrows():
    cliente = row.get('cliente', '')
    codigo = row.get('codigo', '')
    telefone = row.get('telefone', '')
    
    # Parse dias de recebimento
    dias_recebimento = parse_list(row.get('dias_recebimento', '[]'))
    dias_consumo = parse_list(row.get('dias_consumo', '[]'))
    periodo_recebimento = parse_list(row.get('periodo_recebimento', '[]'))
    veiculos_permitidos = parse_list(row.get('veiculos_permitidos', '[]'))
    necessidade_reservar_vaga = parse_list(row.get('necessidade_reservar_vaga', '[]'))
    visibilidade_trajeto = parse_list(row.get('visibilidade_trajeto', '[]'))
    local_facil_acesso = parse_list(row.get('local_facil_acesso', '[]'))
    possui_estacionamento = parse_list(row.get('possui_estacionamento', '[]'))
    
    # Normalizar dias
    dias_recebimento_norm = [normalizar_dia(d) for d in dias_recebimento]
    dias_consumo_norm = [normalizar_dia(d) for d in dias_consumo]
    
    # Extrair primeiro valor (se for lista) para os campos que são listas
    periodo_rec = periodo_recebimento
    necessidade_vaga = necessidade_reservar_vaga[0] if necessidade_reservar_vaga else ''
    visibilidade = visibilidade_trajeto[0] if visibilidade_trajeto else ''
    facil_acesso = local_facil_acesso[0] if local_facil_acesso else ''
    estacionamento = possui_estacionamento[0] if possui_estacionamento else ''
    
    print(f"\n  Cliente: {cliente}")
    print(f"  Código: {codigo}")
    print(f"  Dias de recebimento: {dias_recebimento_norm}")
    print(f"  Dias de consumo: {dias_consumo_norm}")
    print(f"  Período recebimento: {periodo_rec}")
    print(f"  Veículos permitidos: {veiculos_permitidos}")
    print(f"  Necessidade vaga: {necessidade_vaga}")
    print(f"  Visibilidade trajeto: {visibilidade}")
    print(f"  Fácil acesso: {facil_acesso}")
    print(f"  Estacionamento: {estacionamento}")
    
    resultados.append({
        'cliente': cliente,
        'codigo': codigo,
        'telefone': telefone,
        'dias_recebimento': dias_recebimento_norm,
        'dias_consumo': dias_consumo_norm,
        'periodo_recebimento': periodo_rec,
        'veiculos_permitidos': veiculos_permitidos,
        'necessidade_vaga': necessidade_vaga,
        'visibilidade_trajeto': visibilidade,
        'facil_acesso': facil_acesso,
        'estacionamento': estacionamento,
    })

# Agora vamos preencher o modelo 1
print("\n\n🔄 Transformando modelo.xlsx...")

# Carregar com openpyxl para manipulação completa
wb = openpyxl.load_workbook(path_atual / 'modelo.xlsx')
ws = wb.active

# Mapeamento de abreviaturas para dias completos
abrev_para_dia = {
    'Seg': 'Segunda',
    'Ter': 'Terça',
    'Qua': 'Quarta',
    'Qui': 'Quinta',
    'Sex': 'Sexta',
    'Sáb': 'Sábado',
    'Dom': 'Domingo'
}

# Colunas dos dias de recebimento (coluna 10-16)
cols_dias_recebimento = {
    'Segunda': 10,
    'Terça': 11,
    'Quarta': 12,
    'Quinta': 13,
    'Sexta': 14,
    'Sábado': 15,
    'Domingo': 16
}

# Colunas dos dias de consumo (coluna 3-9) - opcional
cols_dias_consumo = {
    'Segunda': 3,
    'Terça': 4,
    'Quarta': 5,
    'Quinta': 6,
    'Sexta': 7,
    'Sábado': 8,
    'Domingo': 9
}

print(f"Colunas dos dias de recebimento: {cols_dias_recebimento}")
print(f"Colunas dos dias de consumo: {cols_dias_consumo}")

# Preencher os dados
linha_dados = 3  # Começar a preenchimento na linha 3

# Mapeamento de períodos de recebimento (coluna 17-21)
cols_periodo = {
    'Manhã': 17,
    'Tarde': 18,
    'Noite': 19,
    'Todos': 20,
    'Hor. comercial': 21,
    'Hor. Comercial': 21,
}

# Mapeamento de veículos permitidos (coluna 22-24)
cols_veiculos = {
    'Toco': 22,
    'toco': 22,
    'Truck': 23,
    'truck': 23,
    'Centopeia': 24,
    'centopeia': 24,
}

# Colunas de sim/não (as últimas 4)
col_necessidade_vaga = 25      # Necessidade_vaga
col_visibilidade = 26           # Visibilidade_trajeto_mangueira
col_facil_acesso = 27           # Fácil_acesso
col_estacionamento = 28         # Estacionamento_interno

for info_cliente in resultados:
    # Preencher nome do cliente (coluna 1)
    ws.cell(row=linha_dados, column=1).value = info_cliente['cliente']
    
    # Preencher código (coluna 2)
    ws.cell(row=linha_dados, column=2).value = str(info_cliente['codigo']) if info_cliente['codigo'] else ''
    
    # Preencher telefone (coluna 29)
    ws.cell(row=linha_dados, column=29).value = info_cliente['telefone'] if not pd.isna(info_cliente['telefone']) else ''
    
    # Preencher dias de consumo (colunas 3-9)
    for dia, col in cols_dias_consumo.items():
        if dia in info_cliente['dias_consumo']:
            ws.cell(row=linha_dados, column=col).value = 'SIM'
        else:
            ws.cell(row=linha_dados, column=col).value = ''
    
    # Preencher dias de recebimento (colunas 10-16)
    for dia, col in cols_dias_recebimento.items():
        if dia in info_cliente['dias_recebimento']:
            ws.cell(row=linha_dados, column=col).value = 'SIM'
        else:
            ws.cell(row=linha_dados, column=col).value = ''
    
    # Preencher período de recebimento (colunas 17-21)
    for periodo in info_cliente['periodo_recebimento']:
        col_periodo = cols_periodo.get(periodo)
        if col_periodo:
            ws.cell(row=linha_dados, column=col_periodo).value = 'SIM'
    
    # Preencher veículos permitidos (colunas 22-24)
    for veiculo in info_cliente['veiculos_permitidos']:
        col_veiculo = cols_veiculos.get(veiculo)
        if col_veiculo:
            ws.cell(row=linha_dados, column=col_veiculo).value = 'SIM'
    
    # Preencher campos SIM/NÃO (últimos 4 campos)
    ws.cell(row=linha_dados, column=col_necessidade_vaga).value = info_cliente['necessidade_vaga']
    ws.cell(row=linha_dados, column=col_visibilidade).value = info_cliente['visibilidade_trajeto']
    ws.cell(row=linha_dados, column=col_facil_acesso).value = info_cliente['facil_acesso']
    ws.cell(row=linha_dados, column=col_estacionamento).value = info_cliente['estacionamento']
    
    linha_dados += 1

# Salvar o arquivo
output_path = path_atual / 'output' / 'modelo_preenchido.xlsx'
wb.save(output_path)

print(f"\n✨ Arquivo salvo com sucesso em: {output_path}")
print("\n✅ Transformação concluída!")